"""
SRA Schedule Intelligence Agent — Comprehensive Bulk Test & Validation Suite
=============================================================================
Tests all 17 tools (16 structured + 1 SQL fallback) across every Layer,
parameter permutation, domain filter, edge case, and representative natural-
language agent query.  Results are written to a multi-sheet Excel workbook.

Architecture
------------
  Layer 0  → sra_sql_agent            (SQL Fallback — 10 patterns)
  Layer 1  → sra_status_pei           (Project Health)
             sra_activity_health      (Activity Drill-Down)
             sra_task_lookahead       (2W / 4W Look-Ahead)
  Layer 2  → sra_critical_path        (CP Analysis)
             sra_float_analysis       (Float & Erosion Watch)
             sra_productivity_rate    (Rate Gap Analysis)
  Layer 3  → sra_baseline_management  (Baseline Drift)
             sra_schedule_quality     (Data Freshness / LAC)
  Layer 4  → sra_drill_delay          (Delay Root Cause)
             sra_procurement_alignment (PRC Rules)
             sra_eot_exposure         (EOT / Contractual Buffer)
  Layer 5  → sra_recovery_advise      (Recovery Options)
             sra_simulate             (What-If Scenarios)
             sra_create_action        (Log Actions)
  Layer 6  → sra_accountability_trace  (Report Flags / Attribution)
             sra_explain_formula      (KPI Explanations)

Test Categories (Sheet Tabs in Excel)
---------------------------------------
  1. SUMMARY      — Pass / Fail / Error rollup per tool & category
  2. LAYER_1      — Schedule Health & Status (Project + Activity + Task)
  3. LAYER_2      — Schedule Intelligence (CP, Float, Productivity)
  4. LAYER_3      — Schedule Integrity (Baseline, Quality)
  5. LAYER_4      — Schedule Risk & Impact (Delay, PRC, EOT)
  6. LAYER_5      — Action & Recovery (Advise, Simulate, Create)
  7. LAYER_6      — Communication & Governance (Trace, Formulas)
  8. SQL_AGENT    — SQL Fallback (all NL patterns + raw SQL + edge cases)
  9. AGENT_NL     — Full graph NL query tests (if --agent flag used)
  10. EDGE_CASES  — Missing inputs, invalid keys, empty filters

Usage
-----
  python -m agent.bulk_test_tools                         # tool-only, project 101
  python -m agent.bulk_test_tools --project 202
  python -m agent.bulk_test_tools --agent                 # also run NL agent graph
  python -m agent.bulk_test_tools --output report.xlsx
  python -m agent.bulk_test_tools --fast                  # skip slow simulate/create
  python -m agent.bulk_test_tools --layer 4               # run only Layer 4 tests
  python -m agent.bulk_test_tools --sql-only              # run only SQL fallback tests

Requires: openpyxl, pandas  (pip install openpyxl pandas)
"""

import asyncio
import argparse
import sys
import os
import re
import traceback
from copy import deepcopy
from datetime import datetime
from typing import Any, Optional

# ── Project root on path ──────────────────────────────────────────────────────
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd

from db import get_prisma, close_prisma
from agent.tools import (
    SRA_TOOLS,
    sra_status_pei,
    sra_activity_health,
    sra_task_lookahead,
    sra_critical_path,
    sra_float_analysis,
    sra_productivity_rate,
    sra_baseline_management,
    sra_schedule_quality,
    sra_drill_delay,
    sra_procurement_alignment,
    sra_eot_exposure,
    sra_recovery_advise,
    sra_simulate,
    sra_create_action,
    sra_accountability_trace,
    sra_explain_formula,
    sra_sql_agent,
)

# ── Tool registry ─────────────────────────────────────────────────────────────
TOOL_BY_NAME: dict[str, Any] = {t.name: t for t in SRA_TOOLS}

# ── Validation helpers ────────────────────────────────────────────────────────
# Keywords that must appear in a valid (non-error) tool response
_EXPECTED_KEYWORDS: dict[str, list[str]] = {
    "sra_status_pei":            ["SPI", "PEI", "Schedule Health", "Programme"],
    "sra_activity_health":       ["Activity", "Status", "Domain", "Workfront"],
    "sra_task_lookahead":        ["Look-Ahead", "Task", "Window", "Forecast"],
    "sra_critical_path":         ["Critical Path", "Float", "CP"],
    "sra_float_analysis":        ["Float", "Near-Critical", "Erosion"],
    "sra_productivity_rate":     ["Rate", "Progress", "Earned", "Planned"],
    "sra_baseline_management":   ["Baseline", "Drift", "JCR", "Slip"],
    "sra_schedule_quality":      ["LAC", "Compliance", "Data Freshness"],
    "sra_drill_delay":           ["Delay", "Hotspot", "Workfront"],
    "sra_procurement_alignment": ["Procurement", "PRC", "LAC"],
    "sra_eot_exposure":          ["EOT", "Contractual", "Buffer"],
    "sra_recovery_advise":       ["Recovery", "Option", "Workfront"],
    "sra_simulate":              ["Scenario", "Projected", "Programme Impact"],
    "sra_create_action":         ["Action Logged", "ACT-", "Recovery Action"],
    "sra_accountability_trace":  ["Accountability", "Report", "Task"],
    "sra_explain_formula":       ["Formula", "Definition", "Interpretation"],
    "sra_sql_agent":             ["Project", "Activity", "Task"],
}

# Keywords that always signal an error response even when no exception raised
_ERROR_SIGNALS = [
    "No schedule data found",
    "Invalid project key",
    "Tool not found",
    "Error retrieving",
    "Error generating",
    "Error running",
    "Database Connection Unavailable",
]


def _validate_output(tool_name: str, output: str) -> tuple[str, str]:
    """
    Returns (status, validation_note).
    status: PASS | WARN | FAIL
    """
    if not output:
        return "FAIL", "Empty output"

    # Check for error signal strings
    for sig in _ERROR_SIGNALS:
        if sig.lower() in output.lower():
            return "WARN", f"Soft error in output: '{sig}'"

    # Check expected keywords
    expected = _EXPECTED_KEYWORDS.get(tool_name, [])
    missing = [kw for kw in expected if kw.lower() not in output.lower()]
    if missing:
        return "WARN", f"Expected keywords missing: {missing}"

    return "PASS", "All expected content present"


def _truncate(text: str, max_len: int = 6000) -> str:
    if not text or not isinstance(text, str):
        return str(text) if text is not None else ""
    return text[:max_len] + f"\n[... truncated at {max_len} chars]" if len(text) > max_len else text


def _row(
    run_id: int,
    layer: str,
    category: str,
    test_name: str,
    tool_name: str,
    input_summary: str,
    kwargs: dict,
    output: str,
    error: str,
    duration_sec: float,
    status: str,
    validation_note: str,
    insight: str = "",
    mode: str = "tool",
) -> dict:
    return {
        "run_id":          run_id,
        "mode":            mode,
        "layer":           layer,
        "category":        category,
        "test_name":       test_name,
        "tool_name":       tool_name,
        "input_summary":   input_summary,
        "kwargs":          str(kwargs),
        "tool_output":     _truncate(output),
        "insight":         _truncate(insight),
        "error":           error,
        "duration_sec":    round(duration_sec, 2),
        "status":          status,
        "validation_note": validation_note,
        "timestamp":       datetime.now().isoformat(),
    }


# ═════════════════════════════════════════════════════════════════════════════
#  MASTER TEST CASE CATALOGUE
#  Each entry: (layer, category, test_name, tool_name, kwargs)
# ═════════════════════════════════════════════════════════════════════════════

# ── placeholder — replaced at runtime with actual project_key ─────────────────
PK = "101"

def build_test_cases(project_key: str) -> list[tuple]:
    """
    Returns the full test case list with `project_key` injected.
    Format: (layer, category, test_name, tool_name, kwargs_dict)
    """
    pk = project_key  # shorthand

    cases: list[tuple[str, str, str, str, dict]] = [

        # ══════════════════════════════════════════════════════════════════════
        #  LAYER 1 — SCHEDULE HEALTH & STATUS
        # ══════════════════════════════════════════════════════════════════════

        # ── sra_status_pei ────────────────────────────────────────────────────
        ("L1", "Health",   "PEI — Full project health snapshot",
         "sra_status_pei", {"project_key": pk}),

        ("L1", "Health",   "PEI — Missing project key (should list projects)",
         "sra_status_pei", {}),

        ("L1", "Health",   "PEI — Invalid project key (non-numeric)",
         "sra_status_pei", {"project_key": "ABC"}),

        ("L1", "Health",   "PEI — Non-existent project key",
         "sra_status_pei", {"project_key": "9999"}),

        # ── sra_activity_health ───────────────────────────────────────────────
        ("L1", "Activity", "Activity — All activities, all domains",
         "sra_activity_health", {"project_key": pk}),

        ("L1", "Activity", "Activity — Engineering domain filter (ENG)",
         "sra_activity_health", {"project_key": pk, "domain_code": "ENG"}),

        ("L1", "Activity", "Activity — Procurement domain filter (PRC)",
         "sra_activity_health", {"project_key": pk, "domain_code": "PRC"}),

        ("L1", "Activity", "Activity — Construction domain filter (CON)",
         "sra_activity_health", {"project_key": pk, "domain_code": "CON"}),

        ("L1", "Activity", "Activity — No project key",
         "sra_activity_health", {}),

        # ── sra_task_lookahead ────────────────────────────────────────────────
        ("L1", "LookAhead", "Task 2W — All domains, 2-week window",
         "sra_task_lookahead", {"project_key": pk, "window": "2W"}),

        ("L1", "LookAhead", "Task 4W — All domains, 4-week window",
         "sra_task_lookahead", {"project_key": pk, "window": "4W"}),

        ("L1", "LookAhead", "Task 2W — Construction domain only",
         "sra_task_lookahead", {"project_key": pk, "window": "2W", "domain_code": "CON"}),

        ("L1", "LookAhead", "Task 4W — Engineering domain only",
         "sra_task_lookahead", {"project_key": pk, "window": "4W", "domain_code": "ENG"}),

        ("L1", "LookAhead", "Task 2W — Procurement domain only",
         "sra_task_lookahead", {"project_key": pk, "window": "2W", "domain_code": "PRC"}),

        ("L1", "LookAhead", "Task 2W — No project key",
         "sra_task_lookahead", {}),

        # ══════════════════════════════════════════════════════════════════════
        #  LAYER 2 — SCHEDULE INTELLIGENCE
        # ══════════════════════════════════════════════════════════════════════

        # ── sra_critical_path ─────────────────────────────────────────────────
        ("L2", "CritPath",  "Critical Path — Full project CP analysis",
         "sra_critical_path", {"project_key": pk}),

        ("L2", "CritPath",  "Critical Path — No project key",
         "sra_critical_path", {}),

        ("L2", "CritPath",  "Critical Path — Non-existent project",
         "sra_critical_path", {"project_key": "9999"}),

        # ── sra_float_analysis ────────────────────────────────────────────────
        ("L2", "Float",    "Float — Full programme float analysis",
         "sra_float_analysis", {"project_key": pk}),

        ("L2", "Float",    "Float — Construction domain float watch",
         "sra_float_analysis", {"project_key": pk, "domain_code": "CON"}),

        ("L2", "Float",    "Float — Engineering domain float watch",
         "sra_float_analysis", {"project_key": pk, "domain_code": "ENG"}),

        ("L2", "Float",    "Float — Procurement domain float watch",
         "sra_float_analysis", {"project_key": pk, "domain_code": "PRC"}),

        ("L2", "Float",    "Float — No project key",
         "sra_float_analysis", {}),

        # ── sra_productivity_rate ─────────────────────────────────────────────
        ("L2", "Rate",     "Productivity — Full rate gap analysis",
         "sra_productivity_rate", {"project_key": pk}),

        ("L2", "Rate",     "Productivity — Construction rate analysis",
         "sra_productivity_rate", {"project_key": pk, "domain_code": "CON"}),

        ("L2", "Rate",     "Productivity — Engineering rate analysis",
         "sra_productivity_rate", {"project_key": pk, "domain_code": "ENG"}),

        ("L2", "Rate",     "Productivity — Procurement rate analysis",
         "sra_productivity_rate", {"project_key": pk, "domain_code": "PRC"}),

        ("L2", "Rate",     "Productivity — No project key",
         "sra_productivity_rate", {}),

        # ══════════════════════════════════════════════════════════════════════
        #  LAYER 3 — SCHEDULE INTEGRITY
        # ══════════════════════════════════════════════════════════════════════

        # ── sra_baseline_management ───────────────────────────────────────────
        ("L3", "Baseline", "Baseline — JCR/PMS dates, drift analysis",
         "sra_baseline_management", {"project_key": pk}),

        ("L3", "Baseline", "Baseline — No project key",
         "sra_baseline_management", {}),

        ("L3", "Baseline", "Baseline — Non-existent project",
         "sra_baseline_management", {"project_key": "9999"}),

        # ── sra_schedule_quality ──────────────────────────────────────────────
        ("L3", "Quality",  "Quality — Data freshness + LAC compliance (full)",
         "sra_schedule_quality", {"project_key": pk}),

        ("L3", "Quality",  "Quality — No project key",
         "sra_schedule_quality", {}),

        ("L3", "Quality",  "Quality — Non-existent project",
         "sra_schedule_quality", {"project_key": "9999"}),

        # ══════════════════════════════════════════════════════════════════════
        #  LAYER 4 — SCHEDULE RISK & IMPACT
        # ══════════════════════════════════════════════════════════════════════

        # ── sra_drill_delay ───────────────────────────────────────────────────
        ("L4", "Delay",   "Delay — Full delay & risk analysis",
         "sra_drill_delay", {"project_key": pk}),

        ("L4", "Delay",   "Delay — With start/end date window",
         "sra_drill_delay", {"project_key": pk,
                              "start_date": "2025-01-01",
                              "end_date":   "2025-06-30"}),

        ("L4", "Delay",   "Delay — Q3 2025 date window",
         "sra_drill_delay", {"project_key": pk,
                              "start_date": "2025-07-01",
                              "end_date":   "2025-09-30"}),

        ("L4", "Delay",   "Delay — No project key",
         "sra_drill_delay", {}),

        ("L4", "Delay",   "Delay — Non-existent project",
         "sra_drill_delay", {"project_key": "9999"}),

        # ── sra_procurement_alignment ─────────────────────────────────────────
        ("L4", "Procurement", "PRC — Procurement rule alignment (full)",
         "sra_procurement_alignment", {"project_key": pk}),

        ("L4", "Procurement", "PRC — No project key",
         "sra_procurement_alignment", {}),

        ("L4", "Procurement", "PRC — Non-existent project",
         "sra_procurement_alignment", {"project_key": "9999"}),

        # ── sra_eot_exposure ──────────────────────────────────────────────────
        ("L4", "EOT",     "EOT — Exposure and contractual buffer analysis",
         "sra_eot_exposure", {"project_key": pk}),

        ("L4", "EOT",     "EOT — No project key",
         "sra_eot_exposure", {}),

        ("L4", "EOT",     "EOT — Non-existent project",
         "sra_eot_exposure", {"project_key": "9999"}),

        # ══════════════════════════════════════════════════════════════════════
        #  LAYER 5 — SCHEDULE ACTION & RECOVERY
        # ══════════════════════════════════════════════════════════════════════

        # ── sra_recovery_advise ───────────────────────────────────────────────
        ("L5", "Recovery", "Recovery — Generic options (no resource type)",
         "sra_recovery_advise", {"project_key": pk}),

        ("L5", "Recovery", "Recovery — Labor-targeted recovery advice",
         "sra_recovery_advise", {"project_key": pk, "resource_type": "labor"}),

        ("L5", "Recovery", "Recovery — Equipment-targeted recovery advice",
         "sra_recovery_advise", {"project_key": pk, "resource_type": "equipment"}),

        ("L5", "Recovery", "Recovery — Material-targeted recovery advice",
         "sra_recovery_advise", {"project_key": pk, "resource_type": "material"}),

        ("L5", "Recovery", "Recovery — Activity-focused (activity_id provided)",
         "sra_recovery_advise", {"project_key": pk,
                                  "resource_type": "labor",
                                  "activity_id":   "ACT-001"}),

        ("L5", "Recovery", "Recovery — No project key",
         "sra_recovery_advise", {}),

        # ── sra_simulate ──────────────────────────────────────────────────────
        ("L5", "Simulate", "Simulate — 2 shuttering gangs",
         "sra_simulate", {"project_key": pk,
                           "resource_type": "shuttering_gang",
                           "value_amount":  2}),

        ("L5", "Simulate", "Simulate — 5 additional labor (manpower)",
         "sra_simulate", {"project_key": pk,
                           "resource_type": "labor",
                           "value_amount":  5}),

        ("L5", "Simulate", "Simulate — 1 excavator deployment",
         "sra_simulate", {"project_key": pk,
                           "resource_type": "equipment",
                           "value_amount":  1}),

        ("L5", "Simulate", "Simulate — Sunday overtime working",
         "sra_simulate", {"project_key": pk,
                           "resource_type": "overtime",
                           "value_amount":  1,
                           "date_range":    "this Sunday"}),

        ("L5", "Simulate", "Simulate — Extended hours with date range",
         "sra_simulate", {"project_key": pk,
                           "resource_type": "extended_hours",
                           "value_amount":  8,
                           "date_range":    "2025-07-15 to 2025-07-20"}),

        ("L5", "Simulate", "Simulate — 3 formwork gangs with date",
         "sra_simulate", {"project_key": pk,
                           "resource_type": "formwork_gang",
                           "value_amount":  3,
                           "date_range":    "next fortnight"}),

        ("L5", "Simulate", "Simulate — Missing resource type (should prompt)",
         "sra_simulate", {"project_key": pk}),

        ("L5", "Simulate", "Simulate — Missing all params (should prompt)",
         "sra_simulate", {}),

        # ── sra_create_action ─────────────────────────────────────────────────
        ("L5", "Actions",  "Create Action — Approve Option 1 (resource augmentation)",
         "sra_create_action", {"project_key": pk,
                                "action_choice": "Approve Option 1 — Resource Augmentation",
                                "user_id":        "site_planner"}),

        ("L5", "Actions",  "Create Action — Approve Option 2 (schedule compression)",
         "sra_create_action", {"project_key": pk,
                                "action_choice": "Approve Option 2 — Schedule Compression",
                                "user_id":        "planning_manager"}),

        ("L5", "Actions",  "Create Action — Fast-track decision",
         "sra_create_action", {"project_key": pk,
                                "action_choice": "Approve Option 3 — Fast-Tracking",
                                "user_id":        "project_director"}),

        ("L5", "Actions",  "Create Action — Raise alert to site planner",
         "sra_create_action", {"project_key": pk,
                                "action_choice": "Raise alert to site planner — CP breach imminent",
                                "user_id":        "site_planner"}),

        ("L5", "Actions",  "Create Action — Escalate to PM (no user_id)",
         "sra_create_action", {"project_key": pk,
                                "action_choice": "Escalate workfront constraints to PM"}),

        ("L5", "Actions",  "Create Action — Notify procurement team",
         "sra_create_action", {"project_key": pk,
                                "action_choice": "Notify procurement team of overdue PRC rules",
                                "user_id":        "procurement_lead"}),

        ("L5", "Actions",  "Create Action — No action choice (should prompt)",
         "sra_create_action", {"project_key": pk}),

        ("L5", "Actions",  "Create Action — No inputs at all",
         "sra_create_action", {}),

        # ══════════════════════════════════════════════════════════════════════
        #  LAYER 6 — SCHEDULE COMMUNICATION & GOVERNANCE
        # ══════════════════════════════════════════════════════════════════════

        # ── sra_accountability_trace ──────────────────────────────────────────
        ("L6", "Accountability", "Trace — All domains, all tasks",
         "sra_accountability_trace", {"project_key": pk}),

        ("L6", "Accountability", "Trace — Construction domain tasks",
         "sra_accountability_trace", {"project_key": pk, "domain_code": "CON"}),

        ("L6", "Accountability", "Trace — Engineering domain tasks",
         "sra_accountability_trace", {"project_key": pk, "domain_code": "ENG"}),

        ("L6", "Accountability", "Trace — Procurement domain tasks",
         "sra_accountability_trace", {"project_key": pk, "domain_code": "PRC"}),

        ("L6", "Accountability", "Trace — No project key",
         "sra_accountability_trace", {}),

        # ── sra_explain_formula ───────────────────────────────────────────────
        ("L6", "Formula",  "Formula — Explain SPI (with live project context)",
         "sra_explain_formula", {"project_key": pk, "metric": "SPI"}),

        ("L6", "Formula",  "Formula — Explain PEI (with live project context)",
         "sra_explain_formula", {"project_key": pk, "metric": "PEI"}),

        ("L6", "Formula",  "Formula — Explain LAC compliance",
         "sra_explain_formula", {"project_key": pk, "metric": "LAC"}),

        ("L6", "Formula",  "Formula — Explain Float",
         "sra_explain_formula", {"project_key": pk, "metric": "float"}),

        ("L6", "Formula",  "Formula — Explain EOT",
         "sra_explain_formula", {"project_key": pk, "metric": "EOT"}),

        ("L6", "Formula",  "Formula — Explain ALL metrics",
         "sra_explain_formula", {"project_key": pk, "metric": "all"}),

        ("L6", "Formula",  "Formula — No project key, SPI only",
         "sra_explain_formula", {"metric": "SPI"}),

        ("L6", "Formula",  "Formula — No inputs at all",
         "sra_explain_formula", {}),

        # ══════════════════════════════════════════════════════════════════════
        #  LAYER 0 — SQL FALLBACK AGENT
        # ══════════════════════════════════════════════════════════════════════

        # ── Natural Language → Auto-synthesised SQL ───────────────────────────
        ("L0", "SQL-NL",   "SQL — NL: List project summary / portfolio view",
         "sra_sql_agent",  {"question": "List all projects with SPI, health score and delay",
                             "project_key": pk}),

        ("L0", "SQL-NL",   "SQL — NL: Critical tasks on CP (project scope)",
         "sra_sql_agent",  {"question": "Show all critical tasks in project",
                             "project_key": pk}),

        ("L0", "SQL-NL",   "SQL — NL: Delayed tasks with risk tier",
         "sra_sql_agent",  {"question": "Show delayed tasks with risk tier",
                             "project_key": pk}),

        ("L0", "SQL-NL",   "SQL — NL: Float distribution across tasks",
         "sra_sql_agent",  {"question": "Float distribution for all tasks",
                             "project_key": pk}),

        ("L0", "SQL-NL",   "SQL — NL: 2-week look-ahead tasks",
         "sra_sql_agent",  {"question": "Show tasks in the 2 week look ahead window",
                             "project_key": pk}),

        ("L0", "SQL-NL",   "SQL — NL: Milestones and their status",
         "sra_sql_agent",  {"question": "List all milestones and their status",
                             "project_key": pk}),

        ("L0", "SQL-NL",   "SQL — NL: Delayed activities with SPI",
         "sra_sql_agent",  {"question": "Which activities have slipped and what is their SPI",
                             "project_key": pk}),

        ("L0", "SQL-NL",   "SQL — NL: Workfront-constrained activities",
         "sra_sql_agent",  {"question": "Activities with low workfront readiness",
                             "project_key": pk}),

        ("L0", "SQL-NL",   "SQL — NL: LAC compliance per activity",
         "sra_sql_agent",  {"question": "Show LAC compliance per activity",
                             "project_key": pk}),

        ("L0", "SQL-NL",   "SQL — NL: Critical activities (CP focus)",
         "sra_sql_agent",  {"question": "Which activities are on the critical path",
                             "project_key": pk}),

        ("L0", "SQL-NL",   "SQL — NL: Procurement at-risk activities",
         "sra_sql_agent",  {"question": "Procurement activities with overdue rules",
                             "project_key": pk}),

        ("L0", "SQL-NL",   "SQL — NL: EOT exposure per project",
         "sra_sql_agent",  {"question": "EOT exposure and contractual buffer for the project",
                             "project_key": pk}),

        ("L0", "SQL-NL",   "SQL — NL: Baseline drift activities",
         "sra_sql_agent",  {"question": "Activities with baseline drift and slip",
                             "project_key": pk}),

        ("L0", "SQL-NL",   "SQL — NL: Portfolio view (no project filter)",
         "sra_sql_agent",  {"question": "Show all projects with health score and SPI overall"}),

        # ── Raw SQL passthrough ───────────────────────────────────────────────
        ("L0", "SQL-RAW",  "SQL RAW — Project summary: key health KPIs",
         "sra_sql_agent",  {
             "question":   "Raw SQL: project KPIs",
             "project_key": pk,
             "raw_sql":    (
                 f'SELECT "Project_Key", "Project_Name", "Project_Location", '
                 f'"SPI_Overall", "Project_Execution_Index", '
                 f'"Max_Forecast_Delay_Days_Overall", "EOT_Exposure_Days", '
                 f'"Schedule_Health_RAG", "Schedule_Health_Score", '
                 f'"Weakest_Domain", "Data_Freshness_Label" '
                 f'FROM public.tbl_01_project_summary '
                 f'WHERE "Project_Key" = {pk} '
                 f'ORDER BY "Max_Forecast_Delay_Days_Overall" DESC NULLS LAST LIMIT 10'
             ),
         }),

        ("L0", "SQL-RAW",  "SQL RAW — Top 10 delayed activities with CP flag",
         "sra_sql_agent",  {
             "question":   "Raw SQL: top delayed activities",
             "project_key": pk,
             "raw_sql":    (
                 f'SELECT "Activity_Code", "Activity_Description", "Domain_Code", '
                 f'"Activity_Status", "Forecast_Delay_Days", "Slip_Days", '
                 f'"Is_Critical_Wrench", "Delay_Magnitude_Label", "Delay_Source_Label", '
                 f'"Compound_Risk_Score" '
                 f'FROM public.tbl_02_project_activity '
                 f'WHERE "Project_Key" = {pk} '
                 f'AND "Forecast_Delay_Days" > 0 '
                 f'ORDER BY "Forecast_Delay_Days" DESC LIMIT 10'
             ),
         }),

        ("L0", "SQL-RAW",  "SQL RAW — Tasks near-critical (float ≤ 5d)",
         "sra_sql_agent",  {
             "question":   "Raw SQL: near-critical tasks",
             "project_key": pk,
             "raw_sql":    (
                 f'SELECT "Task_Key", "Task_Name", "Activity_Code", "Domain_Code", '
                 f'"Total_Float_Days", "Float_Health_Status", "Float_RAG_Label", '
                 f'"Is_Critical_Wrench", "Forecast_Delay_Days", "Task_Health_RAG" '
                 f'FROM public.tbl_03_project_task '
                 f'WHERE "Project_Key" = {pk} '
                 f'AND "Total_Float_Days" >= 0 AND "Total_Float_Days" <= 5 '
                 f'ORDER BY "Total_Float_Days" ASC LIMIT 20'
             ),
         }),

        ("L0", "SQL-RAW",  "SQL RAW — Milestones forecast vs baseline",
         "sra_sql_agent",  {
             "question":   "Raw SQL: milestones",
             "project_key": pk,
             "raw_sql":    (
                 f'SELECT "Task_Key", "Task_Name", "Activity_Code", "Domain_Code", '
                 f'"Baseline_Finish_Date", "Forecast_Finish_Date", '
                 f'"Milestone_Status_Label", "Slip_Days" '
                 f'FROM public.tbl_03_project_task '
                 f'WHERE "Project_Key" = {pk} '
                 f'AND "Is_Milestone" = TRUE '
                 f'ORDER BY "Forecast_Finish_Date" ASC LIMIT 20'
             ),
         }),

        ("L0", "SQL-RAW",  "SQL RAW — Procurement rule summary per activity",
         "sra_sql_agent",  {
             "question":   "Raw SQL: procurement rules",
             "project_key": pk,
             "raw_sql":    (
                 f'SELECT "Activity_Code", "Activity_Description", "Domain_Code", '
                 f'"PRC_Total_Rules", "PRC_Completed_Rules", "PRC_Overdue_Rules", '
                 f'"PRC_At_Risk_Rules_7Day", "PRC_At_Risk_Rules_30Day", '
                 f'"PRC_LAC_Week_Pct", "PRC_LAC_RAG" '
                 f'FROM public.tbl_02_project_activity '
                 f'WHERE "Project_Key" = {pk} '
                 f'AND "PRC_Total_Rules" > 0 '
                 f'ORDER BY "PRC_Overdue_Rules" DESC LIMIT 15'
             ),
         }),

        ("L0", "SQL-RAW",  "SQL RAW — Construction LAC per activity",
         "sra_sql_agent",  {
             "question":   "Raw SQL: construction LAC",
             "project_key": pk,
             "raw_sql":    (
                 f'SELECT "Activity_Code", "Activity_Description", "Domain_Code", '
                 f'"CON_Total_Rules", "CON_Completed_Rules", "CON_Overdue_Rules", '
                 f'"CON_At_Risk_Rules_7Day", "CON_LAC_Week_Pct", "CON_LAC_Month_Pct", '
                 f'"CON_LAC_RAG" '
                 f'FROM public.tbl_02_project_activity '
                 f'WHERE "Project_Key" = {pk} '
                 f'AND "CON_Total_Rules" > 0 '
                 f'ORDER BY "CON_LAC_Week_Pct" ASC LIMIT 15'
             ),
         }),

        ("L0", "SQL-RAW",  "SQL RAW — Recovery candidates (crash + fast-track)",
         "sra_sql_agent",  {
             "question":   "Raw SQL: recovery candidates",
             "project_key": pk,
             "raw_sql":    (
                 f'SELECT "Task_Key", "Task_Name", "Activity_Code", "Domain_Code", '
                 f'"Crash_Candidate_Flag", "FastTrack_Candidate_Flag", '
                 f'"Task_Acceleration_Days", "Task_Recovery_SPI_Required", '
                 f'"Recovery_Feasibility_Label", "Task_Health_RAG" '
                 f'FROM public.tbl_03_project_task '
                 f'WHERE "Project_Key" = {pk} '
                 f'AND ("Crash_Candidate_Flag" = TRUE OR "FastTrack_Candidate_Flag" = TRUE) '
                 f'ORDER BY "Task_Acceleration_Days" DESC NULLS LAST LIMIT 15'
             ),
         }),

        ("L0", "SQL-RAW",  "SQL RAW — As-built vs baseline (completed tasks)",
         "sra_sql_agent",  {
             "question":   "Raw SQL: as-built variance",
             "project_key": pk,
             "raw_sql":    (
                 f'SELECT "Task_Key", "Task_Name", "Activity_Code", "Domain_Code", '
                 f'"Baseline_Duration_Days", "As_Built_Duration_Days", '
                 f'"Duration_Variance_Days", "Task_Status", "Actual_Finish_Date" '
                 f'FROM public.tbl_03_project_task '
                 f'WHERE "Project_Key" = {pk} '
                 f'AND "Task_Status" = \'Complete\' '
                 f'AND "As_Built_Duration_Days" IS NOT NULL '
                 f'ORDER BY "Duration_Variance_Days" DESC NULLS LAST LIMIT 15'
             ),
         }),

        ("L0", "SQL-RAW",  "SQL RAW — Workfront block flag tasks in 2W window",
         "sra_sql_agent",  {
             "question":   "Raw SQL: blocked tasks in 2W window",
             "project_key": pk,
             "raw_sql":    (
                 f'SELECT "Task_Key", "Task_Name", "Activity_Code", "Domain_Code", '
                 f'"Forecast_Start_Date", "Forecast_Finish_Date", '
                 f'"Workfront_Block_Flag", "Is_WorkFront_Available", '
                 f'"Total_Float_Days", "LookAhead_Priority_Score" '
                 f'FROM public.tbl_03_project_task '
                 f'WHERE "Project_Key" = {pk} '
                 f'AND "In_2W_Start_Window" IS NOT NULL '
                 f'ORDER BY "LookAhead_Priority_Score" DESC NULLS LAST LIMIT 20'
             ),
         }),

        # ── SQL Edge & Security Cases ─────────────────────────────────────────
        ("L0", "SQL-EDGE", "SQL EDGE — Blocked: INSERT statement attempt",
         "sra_sql_agent",  {
             "question":   "Security test: INSERT attempt",
             "raw_sql":    "INSERT INTO tbl_01_project_summary VALUES (1)",
         }),

        ("L0", "SQL-EDGE", "SQL EDGE — Blocked: DELETE statement attempt",
         "sra_sql_agent",  {
             "question":   "Security test: DELETE attempt",
             "raw_sql":    "DELETE FROM tbl_01_project_summary WHERE 1=1",
         }),

        ("L0", "SQL-EDGE", "SQL EDGE — Blocked: DROP TABLE attempt",
         "sra_sql_agent",  {
             "question":   "Security test: DROP TABLE attempt",
             "raw_sql":    "DROP TABLE tbl_03_project_task",
         }),

        ("L0", "SQL-EDGE", "SQL EDGE — Blocked: Multiple statements (semicolon)",
         "sra_sql_agent",  {
             "question":   "Security test: multiple statements",
             "raw_sql":    "SELECT 1; DROP TABLE tbl_01_project_summary",
         }),

        ("L0", "SQL-EDGE", "SQL EDGE — Ambiguous question (should request clarification)",
         "sra_sql_agent",  {
             "question":   "Show me everything",
             "project_key": pk,
         }),

        ("L0", "SQL-EDGE", "SQL EDGE — No question, no SQL (empty inputs)",
         "sra_sql_agent",  {
             "question":   "general question with no direction",
         }),

    ]

    return cases


# ═════════════════════════════════════════════════════════════════════════════
#  AGENT NL TEST QUERIES  (--agent mode only)
# ═════════════════════════════════════════════════════════════════════════════

def build_agent_queries(project_key: str) -> list[tuple[str, str]]:
    """
    Returns list of (layer_hint, nl_query) for full agent graph tests.
    """
    pk = project_key
    return [
        # ── L1 Health ─────────────────────────────────────────────────────────
        ("L1", f"What is the overall schedule health of project {pk}?"),
        ("L1", f"Is project {pk} on track? Give me the SPI and PEI."),
        ("L1", f"Give me the activity health summary for project {pk}."),
        ("L1", f"Which Construction activities in project {pk} are behind plan?"),
        ("L1", f"What tasks are starting in the next 2 weeks for project {pk}?"),
        ("L1", f"Show me the 4-week look-ahead for Engineering in project {pk}."),

        # ── L2 Intelligence ───────────────────────────────────────────────────
        ("L2", f"What is driving the critical path in project {pk}?"),
        ("L2", f"How tight is the critical path in project {pk}?"),
        ("L2", f"Show me float erosion watch for project {pk}."),
        ("L2", f"Which activities in project {pk} are near-critical?"),
        ("L2", f"What is the productivity rate gap in Construction for project {pk}?"),
        ("L2", f"Are we achieving the planned progress rate in project {pk}?"),

        # ── L3 Integrity ──────────────────────────────────────────────────────
        ("L3", f"How much has the schedule drifted from the original baseline in project {pk}?"),
        ("L3", f"What is the data freshness and LAC compliance for project {pk}?"),

        # ── L4 Risk ───────────────────────────────────────────────────────────
        ("L4", f"Why is project {pk} delayed?"),
        ("L4", f"Where is the delay hotspot in project {pk}?"),
        ("L4", f"Are procurement milestones aligned with the construction schedule in project {pk}?"),
        ("L4", f"What is our EOT exposure for project {pk}?"),
        ("L4", f"Are we at risk of liquidated damages on project {pk}?"),

        # ── L5 Recovery ───────────────────────────────────────────────────────
        ("L5", f"What are our options to recover the programme for project {pk}?"),
        ("L5", f"What if we deploy 2 shuttering gangs on project {pk}?"),
        ("L5", f"Simulate working this Sunday on project {pk}."),
        ("L5", f"Log a recovery action for project {pk}: Approve Option 2 — Schedule Compression, assign to planning_manager."),

        # ── L6 Governance ─────────────────────────────────────────────────────
        ("L6", f"Which tasks are flagged for reporting this period for project {pk}?"),
        ("L6", "How is SPI calculated?"),
        ("L6", "Explain PEI and what it means for our project duration."),
        ("L6", "What is Float and how does it relate to the critical path?"),
        ("L6", "Explain LAC and why it matters for schedule discipline."),
        ("L6", "What is EOT and when would we be eligible?"),

        # ── L0 SQL direct from NL ─────────────────────────────────────────────
        ("L0", f"Give me a table of all tasks in project {pk} with their float and CP status."),
        ("L0", f"List all milestones in project {pk} sorted by forecast finish date."),
        ("L0", f"Show me a ranked list of activities by forecast delay in project {pk}."),
    ]


# ═════════════════════════════════════════════════════════════════════════════
#  TEST RUNNER
# ═════════════════════════════════════════════════════════════════════════════

async def run_tool_tests(
    project_key: str,
    layer_filter: Optional[str] = None,
    sql_only: bool = False,
    fast: bool = False,
) -> list[dict]:
    """
    Run all structured tool + SQL agent tests.
    Returns list of result row dicts.
    """
    test_cases = build_test_cases(project_key)
    rows: list[dict] = []
    run_id = 1

    # Skip flags
    _SLOW_TOOLS = {"sra_simulate", "sra_create_action"} if fast else set()

    for layer, category, test_name, tool_name, kwargs in test_cases:
        # Apply filters
        if layer_filter and not layer.startswith(layer_filter.upper().lstrip("L")):
            continue
        if sql_only and layer != "L0":
            continue
        if tool_name in _SLOW_TOOLS:
            rows.append(_row(
                run_id, layer, category, test_name, tool_name,
                str(kwargs), kwargs, "", "",
                0, "SKIP", "Skipped in --fast mode",
            ))
            run_id += 1
            continue

        tool = TOOL_BY_NAME.get(tool_name)
        if not tool:
            rows.append(_row(
                run_id, layer, category, test_name, tool_name,
                str(kwargs), kwargs, "", f"Tool not found: {tool_name}",
                0, "FAIL", "Tool missing from SRA_TOOLS registry",
            ))
            run_id += 1
            continue

        start = datetime.now()
        error_msg = ""
        output = ""
        try:
            result = await tool.ainvoke(kwargs)
            if isinstance(result, dict) and "output" in result:
                result = result.get("output", "")
            output = str(result) if result is not None else ""
        except Exception as exc:
            error_msg = f"{type(exc).__name__}: {exc}\n{traceback.format_exc()[-600:]}"

        elapsed = (datetime.now() - start).total_seconds()
        status, note = _validate_output(tool_name, output) if not error_msg else ("FAIL", f"Exception: {error_msg[:120]}")

        rows.append(_row(
            run_id, layer, category, test_name, tool_name,
            test_name, kwargs, output, error_msg,
            elapsed, status, note,
        ))
        run_id += 1

        # Progress print
        icon = "✅" if status == "PASS" else ("⚠️ " if status == "WARN" else "❌")
        print(f"  {icon} [{layer}][{category}] {test_name[:60]:<60} {elapsed:.2f}s")

    return rows


async def run_agent_tests(project_key: str) -> list[dict]:
    """
    Run full LangGraph agent with NL queries; capture tool output + insight.
    """
    try:
        from agent.graph import create_agent
        from langchain_core.messages import HumanMessage, AIMessage, ToolMessage
    except ImportError as e:
        return [_row(1, "AGENT", "Import", "Graph import failed", "N/A",
                     "import", {}, "", str(e), 0, "FAIL", "Import error")]

    queries = build_agent_queries(project_key)
    rows: list[dict] = []
    run_id = 1

    agent = await create_agent(checkpointer=None)
    thread_prefix = f"bulk-agent-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
    print(f"\n[AGENT] Running {len(queries)} NL queries...")

    for layer_hint, query in queries:
        thread_id = f"{thread_prefix}-{run_id}"
        start = datetime.now()
        error_msg = ""
        tool_out = ""
        insight = ""

        try:
            ctx = (
                f"\n\n[CONTEXT]\nSelected Project: {project_key}\n"
                f"When calling tools, use project_key='{project_key}'.\n[/CONTEXT]"
            )
            msg = HumanMessage(content=query + ctx)
            result = await agent.ainvoke(
                {"messages": [msg], "thread_id": thread_id},
                config={"configurable": {"thread_id": thread_id}},
            )
            messages = result.get("messages", [])
            for m in reversed(messages):
                if isinstance(m, ToolMessage) and getattr(m, "content", None):
                    tool_out = m.content or ""
                    break
            for m in reversed(messages):
                if isinstance(m, AIMessage) and getattr(m, "content", None) and not getattr(m, "tool_calls", None):
                    insight = m.content or ""
                    break
        except Exception as exc:
            error_msg = f"{type(exc).__name__}: {exc}\n{traceback.format_exc()[-600:]}"

        elapsed = (datetime.now() - start).total_seconds()
        status = "FAIL" if error_msg else ("PASS" if (tool_out or insight) else "WARN")
        note = "Agent responded" if not error_msg else f"Exception: {error_msg[:100]}"

        rows.append(_row(
            run_id, layer_hint, "NL-Agent",
            query[:80], "agent_graph", query,
            {}, tool_out, error_msg, elapsed,
            status, note, insight, mode="agent",
        ))
        run_id += 1
        icon = "✅" if status == "PASS" else ("⚠️ " if status == "WARN" else "❌")
        print(f"  {icon} [{layer_hint}] {query[:70]:<70} {elapsed:.2f}s")

    return rows


# ═════════════════════════════════════════════════════════════════════════════
#  EXCEL WRITER — multi-sheet workbook
# ═════════════════════════════════════════════════════════════════════════════

# Excel column order per sheet
_FULL_COLS = [
    "run_id", "mode", "layer", "category", "test_name", "tool_name",
    "input_summary", "kwargs", "status", "validation_note",
    "duration_sec", "error", "tool_output", "insight", "timestamp",
]

_SUMMARY_COLS = [
    "layer", "category", "tool_name", "total", "pass", "warn", "fail", "skip",
    "avg_duration_sec", "errors",
]


def _make_summary(rows: list[dict]) -> pd.DataFrame:
    """Rollup pass/fail/warn per tool."""
    from collections import defaultdict
    buckets: dict[tuple, dict] = defaultdict(lambda: {
        "total": 0, "pass": 0, "warn": 0, "fail": 0, "skip": 0,
        "durations": [], "errors": [],
    })
    for r in rows:
        key = (r.get("layer", ""), r.get("category", ""), r.get("tool_name", ""))
        b = buckets[key]
        b["total"] += 1
        s = r.get("status", "").upper()
        if s == "PASS":     b["pass"] += 1
        elif s == "WARN":   b["warn"] += 1
        elif s == "FAIL":   b["fail"] += 1
        elif s == "SKIP":   b["skip"] += 1
        if r.get("duration_sec"):
            b["durations"].append(r["duration_sec"])
        if r.get("error"):
            b["errors"].append(r["error"][:80])

    summary_rows = []
    for (layer, cat, tool), b in sorted(buckets.items()):
        avg_d = round(sum(b["durations"]) / len(b["durations"]), 2) if b["durations"] else 0
        summary_rows.append({
            "layer":           layer,
            "category":        cat,
            "tool_name":       tool,
            "total":           b["total"],
            "pass":            b["pass"],
            "warn":            b["warn"],
            "fail":            b["fail"],
            "skip":            b["skip"],
            "avg_duration_sec": avg_d,
            "errors":          "; ".join(b["errors"])[:300],
        })
    return pd.DataFrame(summary_rows, columns=_SUMMARY_COLS)


def _sheet_rows(rows: list[dict], layer_prefix: str) -> list[dict]:
    return [r for r in rows if r.get("layer", "").startswith(layer_prefix)]


def write_excel(all_rows: list[dict], path: str) -> None:
    """Write multi-sheet Excel workbook with colour-coded status cells."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    base, _ = os.path.splitext(path)
    csv_fallback = base + "_fallback.csv"

    # Colour fills
    green  = PatternFill("solid", fgColor="C6EFCE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    red    = PatternFill("solid", fgColor="FFC7CE")
    blue   = PatternFill("solid", fgColor="BDD7EE")
    grey   = PatternFill("solid", fgColor="D9D9D9")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    bold_font   = Font(bold=True, size=10)
    wrap_align  = Alignment(wrap_text=True, vertical="top")
    thin_side   = Side(style="thin", color="BFBFBF")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def _df(rows_subset: list[dict], cols: list[str]) -> pd.DataFrame:
        df = pd.DataFrame(rows_subset if rows_subset else [{}])
        for c in cols:
            if c not in df.columns:
                df[c] = ""
        return df[[c for c in cols if c in df.columns]]

    def _write_sheet(ws, df: pd.DataFrame, status_col_idx: Optional[int] = None) -> None:
        """Write DataFrame to worksheet with formatting."""
        # Header row
        for ci, col in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=ci, value=col.upper().replace("_", " "))
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        ws.row_dimensions[1].height = 30

        # Data rows
        for ri, record in enumerate(df.itertuples(index=False), 2):
            for ci, value in enumerate(record, 1):
                cell = ws.cell(row=ri, column=ci, value=value)
                cell.alignment = wrap_align
                cell.border = thin_border
                cell.font = Font(size=9)

            # Status colouring
            if status_col_idx:
                status_val = str(ws.cell(row=ri, column=status_col_idx).value or "").upper()
                fill = green if status_val == "PASS" else (
                       yellow if status_val == "WARN" else (
                       grey   if status_val == "SKIP" else
                       red))
                for ci in range(1, len(df.columns) + 1):
                    ws.cell(row=ri, column=ci).fill = fill

        # Column widths
        col_widths = {
            "run_id": 6, "mode": 6, "layer": 5, "category": 14, "test_name": 35,
            "tool_name": 28, "input_summary": 35, "kwargs": 30, "status": 8,
            "validation_note": 35, "duration_sec": 10, "error": 35,
            "tool_output": 80, "insight": 60, "timestamp": 20,
            "total": 7, "pass": 7, "warn": 7, "fail": 7, "skip": 7,
            "avg_duration_sec": 14, "errors": 40,
        }
        for ci, col in enumerate(df.columns, 1):
            w = col_widths.get(col, 18)
            ws.column_dimensions[get_column_letter(ci)].width = w

        ws.freeze_panes = "A2"

    try:
        wb = openpyxl.Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        # ── SUMMARY sheet ─────────────────────────────────────────────────────
        summary_df = _make_summary(all_rows)
        ws_sum = wb.create_sheet("SUMMARY")
        _write_sheet(ws_sum, summary_df)

        # ── LAYER sheets ──────────────────────────────────────────────────────
        layer_sheets = [
            ("LAYER_1_Health",    "L1"),
            ("LAYER_2_Intel",     "L2"),
            ("LAYER_3_Integrity", "L3"),
            ("LAYER_4_Risk",      "L4"),
            ("LAYER_5_Action",    "L5"),
            ("LAYER_6_Govern",    "L6"),
            ("SQL_AGENT",         "L0"),
        ]
        status_col_idx = _FULL_COLS.index("status") + 1

        for sheet_name, layer_prefix in layer_sheets:
            subset = _sheet_rows(all_rows, layer_prefix)
            ws = wb.create_sheet(sheet_name)
            df = _df(subset, _FULL_COLS)
            _write_sheet(ws, df, status_col_idx)

        # ── AGENT_NL sheet ────────────────────────────────────────────────────
        agent_rows = [r for r in all_rows if r.get("mode") == "agent"]
        ws_agent = wb.create_sheet("AGENT_NL")
        df_agent = _df(agent_rows, _FULL_COLS)
        _write_sheet(ws_agent, df_agent, status_col_idx)

        # ── EDGE_CASES sheet ─────────────────────────────────────────────────
        edge_rows = [r for r in all_rows if r.get("category") in ("SQL-EDGE", "EdgeCase")]
        # Also include tests with non-existent project or missing inputs
        edge_rows += [
            r for r in all_rows
            if "Non-existent" in r.get("test_name", "")
            or "No project key" in r.get("test_name", "")
            or "Missing" in r.get("test_name", "")
            or "Invalid" in r.get("test_name", "")
        ]
        # Deduplicate
        seen_ids = set()
        deduped_edge = []
        for r in edge_rows:
            rid = r.get("run_id")
            if rid not in seen_ids:
                seen_ids.add(rid)
                deduped_edge.append(r)

        ws_edge = wb.create_sheet("EDGE_CASES")
        df_edge = _df(deduped_edge, _FULL_COLS)
        _write_sheet(ws_edge, df_edge, status_col_idx)

        # ── ALL_RESULTS sheet (flat full export) ─────────────────────────────
        ws_all = wb.create_sheet("ALL_RESULTS")
        df_all = _df(all_rows, _FULL_COLS)
        _write_sheet(ws_all, df_all, status_col_idx)

        # Re-order sheets so SUMMARY is first
        wb.move_sheet("SUMMARY", offset=-len(wb.sheetnames) + 1)

        wb.save(path)
        print(f"\n✅ Report written: {path}")
        print(f"   Sheets: {', '.join(wb.sheetnames)}")

    except Exception as exc:
        print(f"⚠️  Excel write failed ({exc}); writing CSV fallback...")
        pd.DataFrame(all_rows).to_csv(csv_fallback, index=False)
        print(f"   CSV written: {csv_fallback}")


# ═════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═════════════════════════════════════════════════════════════════════════════

async def main():
    parser = argparse.ArgumentParser(
        description="SRA Comprehensive Bulk Test — all tools, layers, edge cases → Excel"
    )
    parser.add_argument("--output",   "-o", default="sra_bulk_test_report.xlsx",
                        help="Output Excel file path (default: sra_bulk_test_report.xlsx)")
    parser.add_argument("--project",  "-p", default="101",
                        help="Project key for tests (default: 101)")
    parser.add_argument("--agent",    action="store_true",
                        help="Also run full agent graph with NL queries")
    parser.add_argument("--fast",     action="store_true",
                        help="Skip slow tools (sra_simulate, sra_create_action)")
    parser.add_argument("--sql-only", action="store_true",
                        help="Run only SQL fallback agent tests")
    parser.add_argument("--layer",    default=None,
                        help="Run only a specific layer: 0, 1, 2, 3, 4, 5, or 6")
    parser.add_argument("--tool-only", action="store_true",
                        help="Run structured tool tests only (no agent graph)")
    args = parser.parse_args()

    run_agent = args.agent
    run_tool  = not run_agent or args.tool_only or True  # always run tools unless --agent-only

    print("=" * 70)
    print("  SRA SCHEDULE INTELLIGENCE AGENT — BULK TEST SUITE")
    print("=" * 70)
    print(f"  Project Key : {args.project}")
    print(f"  Mode        : {'Tool + Agent' if run_agent else 'Tool-only'}")
    print(f"  Layer filter: {args.layer or 'All'}")
    print(f"  Fast mode   : {args.fast}")
    print(f"  SQL only    : {args.sql_only}")
    print(f"  Output      : {args.output}")
    print("=" * 70)

    # Count planned tests
    test_cases = build_test_cases(args.project)
    if args.sql_only:
        planned = sum(1 for t in test_cases if t[0] == "L0")
    elif args.layer:
        planned = sum(1 for t in test_cases if t[0].endswith(args.layer))
    else:
        planned = len(test_cases)
    print(f"\n  Planned tool tests: {planned}")
    if run_agent:
        print(f"  Planned agent queries: {len(build_agent_queries(args.project))}")
    print()

    # DB connect
    print("Connecting to DB...")
    await get_prisma()
    all_rows: list[dict] = []

    try:
        if run_tool:
            print("\n[TOOL TESTS] Running...\n")
            tool_rows = await run_tool_tests(
                args.project,
                layer_filter=args.layer,
                sql_only=args.sql_only,
                fast=args.fast,
            )
            all_rows.extend(tool_rows)
            pass_c  = sum(1 for r in tool_rows if r["status"] == "PASS")
            warn_c  = sum(1 for r in tool_rows if r["status"] == "WARN")
            fail_c  = sum(1 for r in tool_rows if r["status"] == "FAIL")
            skip_c  = sum(1 for r in tool_rows if r["status"] == "SKIP")
            print(f"\n  Tool results: {len(tool_rows)} total | "
                  f"✅ {pass_c} PASS | ⚠️  {warn_c} WARN | ❌ {fail_c} FAIL | ⏭  {skip_c} SKIP")

        if run_agent:
            agent_rows = await run_agent_tests(args.project)
            all_rows.extend(agent_rows)
            a_pass = sum(1 for r in agent_rows if r["status"] == "PASS")
            a_fail = sum(1 for r in agent_rows if r["status"] == "FAIL")
            print(f"\n  Agent results: {len(agent_rows)} total | ✅ {a_pass} PASS | ❌ {a_fail} FAIL")

    finally:
        await close_prisma()

    # Re-number run_ids sequentially
    for i, r in enumerate(all_rows, 1):
        r["run_id"] = i

    if not all_rows:
        print("No results — nothing to write.")
        return

    write_excel(all_rows, args.output)

    # Final summary to console
    total  = len(all_rows)
    passed = sum(1 for r in all_rows if r["status"] == "PASS")
    warned = sum(1 for r in all_rows if r["status"] == "WARN")
    failed = sum(1 for r in all_rows if r["status"] == "FAIL")
    skipped= sum(1 for r in all_rows if r["status"] == "SKIP")

    print("\n" + "=" * 70)
    print("  FINAL SUMMARY")
    print("=" * 70)
    print(f"  Total Tests : {total}")
    print(f"  ✅ PASS     : {passed}  ({passed/total*100:.0f}%)")
    print(f"  ⚠️  WARN    : {warned}  ({warned/total*100:.0f}%)")
    print(f"  ❌ FAIL     : {failed}  ({failed/total*100:.0f}%)")
    print(f"  ⏭  SKIP    : {skipped}  ({skipped/total*100:.0f}%)")
    print("=" * 70)
    print(f"  Report: {os.path.abspath(args.output)}")
    print("=" * 70)


if __name__ == "__main__":
    asyncio.run(main())
